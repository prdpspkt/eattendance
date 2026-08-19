"""Administrator views: manual attendance corrections and the overtime report."""
from datetime import date, datetime, timedelta
from decimal import Decimal

from django.contrib import messages
from django.contrib.auth.decorators import login_required, permission_required
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone

from core.models import Department, Employee

from .forms import ManualAttendanceForm, OvertimeRecordForm
from .models import Attendance, OvertimeRecord


def _parse_date(value, fallback=None):
    try:
        return datetime.strptime(value, '%Y-%m-%d').date()
    except (TypeError, ValueError):
        return fallback


def _filter_date(request, field, fallback=None):
    """Read a date filter typed in either calendar.

    The forms submit the AD field the browser's picker fills in, and a BS
    field beside it under ``<field>_bs``. Reading both here means the filter
    still works when the script that keeps them in step has not run.
    """
    from core.nepali_date import parse_either

    return parse_either(
        request.GET.get(field), request.GET.get(f'{field}_bs'), fallback
    )


# ---------------------------------------------------------------------------
# Manual attendance
# ---------------------------------------------------------------------------
@login_required
@permission_required('core.can_manage_devices', raise_exception=False, login_url='dashboard')
def manual_attendance(request):
    """Record a punch a terminal never captured, with the reason it is needed."""
    if request.method == 'POST':
        form = ManualAttendanceForm(request.POST, recorded_by=request.user)
        if form.is_valid():
            punch = form.save()
            messages.success(
                request,
                f'Recorded {punch.get_punch_type_display().lower()} for '
                f'{punch.employee.user.get_full_name()} at '
                f'{timezone.localtime(punch.timestamp):%d %b %Y, %H:%M}. '
                f'The day has been recalculated.'
            )
            return redirect('attendance:manual_attendance')
    else:
        form = ManualAttendanceForm(recorded_by=request.user)

    recent = (
        Attendance.objects
        .filter(source=Attendance.SOURCE_MANUAL)
        .select_related('employee__user', 'recorded_by')
        .order_by('-created_at')[:15]
    )
    return render(request, 'attendance/manual_attendance.html', {
        'form': form,
        'recent_manual': recent,
    })


# ---------------------------------------------------------------------------
# Overtime
# ---------------------------------------------------------------------------
def _overtime_queryset(request):
    """Apply the report's filters, and report back what was applied."""
    today = timezone.localdate()
    start = _filter_date(request, 'start', today.replace(day=1))
    end = _filter_date(request, 'end', today)

    records = (
        OvertimeRecord.objects
        .filter(date__gte=start, date__lte=end)
        .select_related('employee__user', 'employee__department', 'approved_by')
        .order_by('date', 'employee__employee_id')
    )

    employee_id = request.GET.get('employee') or ''
    if employee_id.isdigit():
        records = records.filter(employee_id=int(employee_id))

    department_id = request.GET.get('department') or ''
    if department_id.isdigit():
        records = records.filter(employee__department_id=int(department_id))

    status = request.GET.get('status') or ''
    if status in dict(OvertimeRecord.STATUS_CHOICES):
        records = records.filter(status=status)

    return records, {
        'start': start, 'end': end,
        'employee': employee_id, 'department': department_id, 'status': status,
    }


def _overtime_totals(records):
    """Hours and money for the filtered set. Takes an already-evaluated list.

    Both totals are accumulated per record rather than aggregated in SQL. For
    the amount that is a correctness requirement, not a preference: each record
    snapshots its own rate at approval, so rows in one report legitimately
    carry different rates and (total hours x one rate) would misprice them.
    Summing the hours in the same loop then costs nothing and avoids a second
    query over rows already in memory.
    """
    hours = Decimal('0')
    amount = Decimal('0')
    priced = unpriced = 0
    for record in records:
        hours += record.hours or Decimal('0')
        value = record.amount
        if value is None:
            unpriced += 1
        else:
            amount += value
            priced += 1
    return {
        'hours': hours,
        # None, not zero: "nothing is priced yet" and "the total is zero" are
        # different answers, and a payment report must not blur them.
        'amount': amount if priced else None,
        'priced': priced,
        'unpriced': unpriced,
        'count': len(records),
    }


@login_required
@permission_required('core.can_manage_devices', raise_exception=False, login_url='dashboard')
def overtime_report(request):
    """Overtime for a period: hours, when they were worked, and on what."""
    records, filters = _overtime_queryset(request)
    records = list(records)

    return render(request, 'attendance/overtime_report.html', {
        'records': records,
        'filters': filters,
        'totals': _overtime_totals(records),
        'employees': Employee.objects.select_related('user').order_by('user__first_name'),
        'departments': Department.objects.order_by('name'),
        'status_choices': OvertimeRecord.STATUS_CHOICES,
        'undescribed': sum(1 for r in records if r.needs_job_description),
    })


def _sheet_text(value):
    """Text safe to put in a worksheet cell.

    Names arrive from the terminals, and a terminal will happily report a name
    with a control byte in it - EMP0030 here is enrolled as "NN-\\x05". openpyxl
    refuses to write those characters and raises, which turned one malformed
    enrolment into a payroll export that failed for everybody. Strip them:
    the character carries no meaning, and a slightly odd name in a spreadsheet
    is better than no spreadsheet.
    """
    from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE

    if value is None:
        return ''
    return ILLEGAL_CHARACTERS_RE.sub('', str(value))


def _summarise_by_employee(records):
    """One row per employee: their overtime over the filtered period.

    The money is summed per record for the same reason as in
    ``_overtime_totals`` - each approved record carries the rate that applied
    when it was signed off, so an employee whose rate changed mid-period has
    two rates in one row and only a per-record sum prices them correctly.
    """
    summary = {}
    for record in records:
        employee = record.employee
        row = summary.get(employee.id)
        if row is None:
            row = summary[employee.id] = {
                'employee': employee,
                'days': 0,
                'hours': Decimal('0'),
                'approved_hours': Decimal('0'),
                'pending_hours': Decimal('0'),
                'rejected_hours': Decimal('0'),
                'amount': Decimal('0'),
                'priced': 0,
                'unpriced': 0,
                'undescribed': 0,
                'first_date': record.date,
                'last_date': record.date,
            }

        hours = record.hours or Decimal('0')
        row['days'] += 1
        row['hours'] += hours
        if record.status == OvertimeRecord.STATUS_APPROVED:
            row['approved_hours'] += hours
        elif record.status == OvertimeRecord.STATUS_PENDING:
            row['pending_hours'] += hours
        else:
            row['rejected_hours'] += hours

        value = record.amount
        if value is None:
            row['unpriced'] += 1
        else:
            row['amount'] += value
            row['priced'] += 1

        if record.needs_job_description:
            row['undescribed'] += 1
        row['first_date'] = min(row['first_date'], record.date)
        row['last_date'] = max(row['last_date'], record.date)

    rows = sorted(summary.values(), key=lambda row: row['employee'].employee_id)
    for row in rows:
        # Same distinction as the report totals: nothing priced is not zero.
        if not row['priced']:
            row['amount'] = None
    return rows


def _summarise_by_month(records):
    """Month-by-month totals for the single-employee view.

    Grouped by the Bikram Sambat month, because that is the month the office
    works in - "overtime for Shrawan" is the question being asked, and a BS
    month straddles two AD ones. The AD dates each row covers are carried
    alongside so the payroll figure can still be tied back to the calendar the
    database stores.
    """
    from core.nepali_date import MONTH_NAMES, to_bs

    months = {}
    for record in records:
        converted = to_bs(record.date)
        if converted is None:
            # Outside the BS table: keep the row rather than drop the hours,
            # and label it with the AD month it belongs to.
            key = (record.date.year, record.date.month, 'AD')
            label = record.date.strftime('%B %Y')
        else:
            key = (converted.year, converted.month, 'BS')
            label = f'{MONTH_NAMES[converted.month - 1]} {converted.year}'

        row = months.get(key)
        if row is None:
            row = months[key] = {
                'label': label,
                'days': 0,
                'hours': Decimal('0'),
                'amount': Decimal('0'),
                'priced': 0,
                'first_date': record.date,
                'last_date': record.date,
            }
        row['days'] += 1
        row['hours'] += record.hours or Decimal('0')
        row['first_date'] = min(row['first_date'], record.date)
        row['last_date'] = max(row['last_date'], record.date)
        value = record.amount
        if value is not None:
            row['amount'] += value
            row['priced'] += 1

    rows = [months[key] for key in sorted(months)]
    for row in rows:
        if not row['priced']:
            row['amount'] = None
    return rows


@login_required
@permission_required('core.can_manage_devices', raise_exception=False, login_url='dashboard')
def overtime_summary(request):
    """Overtime totalled over a period - per employee, or for one employee.

    The report next door lists every record; this answers the question payroll
    actually asks, which is how much each person is owed for the month. Same
    filters, so the two are two views of one selection rather than two reports
    that can disagree.
    """
    records, filters = _overtime_queryset(request)
    records = list(records)

    rows = _summarise_by_employee(records)

    # One employee selected: show how the total is made up rather than a
    # single row that says the same thing as the grand total.
    single = rows[0] if (filters['employee'].isdigit() and len(rows) == 1) else None
    months = _summarise_by_month(records) if single else []

    return render(request, 'attendance/overtime_summary.html', {
        'rows': rows,
        'months': months,
        'single': single,
        'records': records if single else [],
        'filters': filters,
        'totals': _overtime_totals(records),
        'employees': Employee.objects.select_related('user').order_by('user__first_name'),
        'departments': Department.objects.order_by('name'),
        'status_choices': OvertimeRecord.STATUS_CHOICES,
    })


@login_required
@permission_required('core.can_manage_devices', raise_exception=False, login_url='dashboard')
def overtime_summary_export(request):
    """The summary as an Excel sheet: one row per employee, plus a total."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    records, filters = _overtime_queryset(request)
    records = list(records)
    rows = _summarise_by_employee(records)
    totals = _overtime_totals(records)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Overtime summary'

    sheet['A1'] = 'Overtime summary'
    sheet['A1'].font = Font(size=14, bold=True)
    sheet['A2'] = f"{filters['start']:%d %b %Y} to {filters['end']:%d %b %Y}"
    sheet['A2'].font = Font(italic=True, color='666666')

    headers = ['Employee ID', 'Employee', 'Department', 'Days', 'Total hours',
               'Approved hours', 'Pending hours', 'Amount', 'Unpriced records']
    header_row = 4
    for column, title in enumerate(headers, start=1):
        cell = sheet.cell(row=header_row, column=column, value=title)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', start_color='1E40AF')
        cell.alignment = Alignment(vertical='center')

    for offset, row in enumerate(rows, start=header_row + 1):
        employee = row['employee']
        sheet.cell(row=offset, column=1, value=_sheet_text(employee.employee_id))
        sheet.cell(row=offset, column=2, value=_sheet_text(employee.user.get_full_name()))
        sheet.cell(row=offset, column=3,
                   value=_sheet_text(employee.department.name) if employee.department else '')
        sheet.cell(row=offset, column=4, value=row['days'])
        sheet.cell(row=offset, column=5, value=float(row['hours']))
        sheet.cell(row=offset, column=6, value=float(row['approved_hours']))
        sheet.cell(row=offset, column=7, value=float(row['pending_hours']))
        sheet.cell(row=offset, column=8,
                   value=float(row['amount']) if row['amount'] is not None else '')
        sheet.cell(row=offset, column=9, value=row['unpriced'] or '')

    total_row = header_row + len(rows) + 1
    sheet.cell(row=total_row, column=4, value='Total').font = Font(bold=True)
    sheet.cell(row=total_row, column=5, value=float(totals['hours'])).font = Font(bold=True)
    if totals['amount'] is not None:
        sheet.cell(row=total_row, column=8, value=float(totals['amount'])).font = Font(bold=True)
    if totals['unpriced']:
        sheet.cell(
            row=total_row + 1, column=1,
            value=f"{totals['unpriced']} record(s) have no rate set and are excluded from the total.",
        ).font = Font(italic=True, color='92400E')

    widths = [12, 26, 18, 8, 12, 15, 14, 14, 16]
    for column, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(column)].width = width

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"overtime-summary-{filters['start']:%Y%m%d}-{filters['end']:%Y%m%d}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    workbook.save(response)
    return response


@login_required
@permission_required('core.can_manage_devices', raise_exception=False, login_url='dashboard')
def overtime_edit(request, record_id):
    """Write down what was worked on, and set the rate."""
    record = get_object_or_404(
        OvertimeRecord.objects.select_related('employee__user'), pk=record_id,
    )

    if request.method == 'POST':
        form = OvertimeRecordForm(request.POST, instance=record)
        if form.is_valid():
            form.save()
            messages.success(request, 'Overtime record updated.')
            return redirect('attendance:overtime_report')
    else:
        form = OvertimeRecordForm(instance=record)

    return render(request, 'attendance/overtime_form.html', {'form': form, 'record': record})


@login_required
@permission_required('core.can_manage_devices', raise_exception=False, login_url='dashboard')
def overtime_approve(request, record_id):
    """Sign a record off. Refuses if nobody has said what the work was."""
    record = get_object_or_404(OvertimeRecord, pk=record_id)

    if request.method != 'POST':
        return redirect('attendance:overtime_report')

    if record.needs_job_description:
        messages.error(
            request,
            'Add the job performed before approving. Overtime is paid on the strength '
            'of the work described.'
        )
    else:
        record.approve(request.user)
        messages.success(
            request,
            f'Approved {record.hours}h for {record.employee.user.get_full_name()} '
            f'on {record.date:%d %b %Y}. The hours are now fixed.'
        )
    return redirect(f"{request.META.get('HTTP_REFERER') or '/attendance/overtime/'}")


@login_required
@permission_required('core.can_manage_devices', raise_exception=False, login_url='dashboard')
def overtime_export(request):
    """The same filtered report as an Excel sheet, for payroll."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    records, filters = _overtime_queryset(request)
    records = list(records)
    totals = _overtime_totals(records)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Overtime'

    sheet['A1'] = 'Overtime report'
    sheet['A1'].font = Font(size=14, bold=True)
    sheet['A2'] = f"{filters['start']:%d %b %Y} to {filters['end']:%d %b %Y}"
    sheet['A2'].font = Font(italic=True, color='666666')

    headers = ['Employee ID', 'Employee', 'Department', 'Date', 'Started', 'Ended',
               'Hours', 'Job performed', 'Rate', 'Amount', 'Status', 'Approved by']
    header_row = 4
    for column, title in enumerate(headers, start=1):
        cell = sheet.cell(row=header_row, column=column, value=title)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', start_color='1E40AF')
        cell.alignment = Alignment(vertical='center')

    for offset, record in enumerate(records, start=header_row + 1):
        local_start = timezone.localtime(record.start_at)
        local_end = timezone.localtime(record.end_at)
        sheet.cell(row=offset, column=1, value=_sheet_text(record.employee.employee_id))
        sheet.cell(row=offset, column=2, value=_sheet_text(record.employee.user.get_full_name()))
        sheet.cell(row=offset, column=3,
                   value=_sheet_text(record.employee.department.name) if record.employee.department else '')
        sheet.cell(row=offset, column=4, value=record.date.strftime('%Y-%m-%d'))
        sheet.cell(row=offset, column=5, value=local_start.strftime('%H:%M'))
        sheet.cell(row=offset, column=6, value=local_end.strftime('%H:%M'))
        sheet.cell(row=offset, column=7, value=float(record.hours))
        # Left blank rather than filled with a placeholder: an empty cell on a
        # payment sheet is a question someone has to answer, which is correct.
        sheet.cell(row=offset, column=8, value=_sheet_text(record.job_performed))
        sheet.cell(row=offset, column=9,
                   value=float(record.hourly_rate) if record.hourly_rate is not None else '')
        amount = record.amount
        sheet.cell(row=offset, column=10, value=float(amount) if amount is not None else '')
        sheet.cell(row=offset, column=11, value=record.get_status_display())
        sheet.cell(row=offset, column=12,
                   value=_sheet_text(record.approved_by.get_full_name()) if record.approved_by else '')

    total_row = header_row + len(records) + 1
    sheet.cell(row=total_row, column=6, value='Total').font = Font(bold=True)
    sheet.cell(row=total_row, column=7, value=float(totals['hours'])).font = Font(bold=True)
    if totals['amount'] is not None:
        sheet.cell(row=total_row, column=10, value=float(totals['amount'])).font = Font(bold=True)
    if totals['unpriced']:
        sheet.cell(
            row=total_row + 1, column=1,
            value=f"{totals['unpriced']} record(s) have no rate set and are excluded from the total.",
        ).font = Font(italic=True, color='92400E')

    widths = [12, 24, 18, 12, 10, 10, 8, 46, 10, 12, 12, 20]
    for column, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(column)].width = width

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"overtime-{filters['start']:%Y%m%d}-{filters['end']:%Y%m%d}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    workbook.save(response)
    return response
